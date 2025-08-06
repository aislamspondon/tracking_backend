
import os
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "valley_hatchery_tracking.settings")  # Change to your settings path
django.setup()

from openpyxl import load_workbook

from tracking_api.models import Tracking

# Load the Excel workbook
wb = load_workbook("trackings.xlsx")
ws = wb.active  # or wb["Trackings"] if you want to specify sheet name

# Iterate over rows (skip header row)
tracking_list = Tracking.objects.all()
# print("Tracking List:", tracking_list)
for tracking in tracking_list:
    print("Tracking Number in Model:", tracking.tracking_number)
    model_tracking_number = tracking.tracking_number
    for row in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 skips the header
        id_, tracking_number = row
        if model_tracking_number == tracking_number:
            print(f"Found matching tracking number: {tracking_number} with ID: {id_}")
            tracking.tracking_id = id_
            tracking.save()

